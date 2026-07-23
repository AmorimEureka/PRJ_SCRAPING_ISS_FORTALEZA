from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from dataclasses import asdict, dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


REQUIRED_COLUMNS = {
    "PACIENTE",
    "LOCAL",
    "TIPO ATENDIMENTO",
    "CPF",
    "VALOR",
    "RUA",
    "NUMERO CASA",
    "BAIRRO",
    "CIDADE",
    "UF",
    "TIPO DE EXAME",
    "EMAIL",
}


@dataclass(frozen=True)
class InvoiceSpreadsheetRow:
    row_number: int
    paciente: str
    local: str
    tipo_atendimento: str
    atendimento: str
    cpf: str
    valor: Decimal
    rua: str
    numero_casa: str
    bairro: str
    cidade: str
    uf: str
    tipo_exame: str
    data: str
    email: str

    @property
    def cpf_digits(self) -> str:
        return re.sub(r"\D", "", self.cpf)

    @property
    def cpf_formatted(self) -> str:
        digits = self.cpf_digits
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"

    @property
    def valor_br(self) -> str:
        return f"{self.valor:.2f}".replace(".", ",")

    @property
    def row_hash(self) -> str:
        values = asdict(self)
        values["valor"] = self.valor_br
        canonical = json.dumps(values, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        return hashlib.sha256(canonical.encode("utf-8")).hexdigest()

    def pdf_filename(self, numero_nfse: str) -> str:
        stem = (
            f"{_safe_filename_part(numero_nfse)} - "
            f"{_safe_filename_part(self.local)}, "
            f"{_safe_filename_part(self.tipo_atendimento)} e "
            f"{_safe_filename_part(self.paciente)}"
        )
        return f"{stem[:230].rstrip()}.pdf"


def load_invoice_rows(path: Path) -> list[InvoiceSpreadsheetRow]:
    if not path.is_file():
        raise FileNotFoundError(f"Planilha nao encontrada: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = {_normalize_header(value): index for index, value in enumerate(raw_headers) if value is not None}
        missing = sorted(REQUIRED_COLUMNS - set(headers))
        if missing:
            raise ValueError("Colunas obrigatorias ausentes na planilha: " + ", ".join(missing))

        rows: list[InvoiceSpreadsheetRow] = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in values):
                continue
            row = _build_row(row_number, values, headers)
            rows.append(row)
        return rows
    finally:
        workbook.close()


def select_rows(
    rows: Iterable[InvoiceSpreadsheetRow],
    *,
    row_number: int | None,
    all_rows: bool,
    limit: int | None,
) -> list[InvoiceSpreadsheetRow]:
    available = list(rows)
    if row_number is not None and all_rows:
        raise ValueError("Use --linha ou --todas, nao ambos.")
    if row_number is None and not all_rows:
        raise ValueError("Informe --linha para uma emissao ou --todas para processar a planilha.")
    if limit is not None and limit < 1:
        raise ValueError("--limite deve ser maior que zero.")

    if row_number is not None:
        selected = [row for row in available if row.row_number == row_number]
        if not selected:
            raise ValueError(f"Linha {row_number} nao encontrada ou vazia na planilha.")
        _validate_row(selected[0])
        return selected

    selected = available[:limit] if limit is not None else available
    for row in selected:
        _validate_row(row)
    return selected


def is_valid_cpf(value: str) -> bool:
    digits = re.sub(r"\D", "", value)
    if len(digits) != 11 or digits == digits[0] * 11:
        return False

    for length in (9, 10):
        total = sum(int(digit) * weight for digit, weight in zip(digits[:length], range(length + 1, 1, -1)))
        check = (total * 10) % 11
        if check == 10:
            check = 0
        if check != int(digits[length]):
            return False
    return True


def _build_row(
    row_number: int,
    values: tuple[Any, ...],
    headers: dict[str, int],
) -> InvoiceSpreadsheetRow:
    def cell(column: str) -> Any:
        index = headers.get(column)
        return values[index] if index is not None and index < len(values) else None

    return InvoiceSpreadsheetRow(
        row_number=row_number,
        paciente=_cell_text(cell("PACIENTE")),
        local=_cell_text(cell("LOCAL")),
        tipo_atendimento=_cell_text(cell("TIPO ATENDIMENTO")),
        atendimento=_cell_text(cell("ATENDIMENTO")),
        cpf=_cell_text(cell("CPF")),
        valor=_decimal_value(cell("VALOR"), row_number),
        rua=_cell_text(cell("RUA")),
        numero_casa=_cell_text(cell("NUMERO CASA")),
        bairro=_cell_text(cell("BAIRRO")),
        cidade=_cell_text(cell("CIDADE")),
        uf=_cell_text(cell("UF")),
        tipo_exame=_cell_text(cell("TIPO DE EXAME")),
        data=_cell_text(cell("DATA")),
        email=_cell_text(cell("EMAIL")),
    )


def _validate_row(row: InvoiceSpreadsheetRow) -> None:
    required_values = {
        "PACIENTE": row.paciente,
        "LOCAL": row.local,
        "TIPO ATENDIMENTO": row.tipo_atendimento,
        "CPF": row.cpf,
        "CIDADE": row.cidade,
        "UF": row.uf,
        "TIPO DE EXAME": row.tipo_exame,
    }
    missing = [name for name, value in required_values.items() if not value]
    if missing:
        raise ValueError(f"Linha {row.row_number}: campos obrigatorios vazios: {', '.join(missing)}")
    if not is_valid_cpf(row.cpf):
        raise ValueError(f"Linha {row.row_number}: CPF invalido: {row.cpf!r}")
    if row.valor <= 0:
        raise ValueError(f"Linha {row.row_number}: VALOR deve ser maior que zero.")
    if len(row.uf) != 2:
        raise ValueError(f"Linha {row.row_number}: UF invalida: {row.uf!r}")


def _normalize_header(value: Any) -> str:
    text = _cell_text(value).upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def _decimal_value(value: Any, row_number: int) -> Decimal:
    if isinstance(value, Decimal):
        result = value
    elif isinstance(value, (int, float)):
        result = Decimal(str(value))
    else:
        text = _cell_text(value)
        if not text:
            raise ValueError(f"Linha {row_number}: VALOR vazio.")
        normalized = text.replace("R$", "").replace(" ", "")
        if "," in normalized:
            normalized = normalized.replace(".", "").replace(",", ".")
        try:
            result = Decimal(normalized)
        except InvalidOperation as exc:
            raise ValueError(f"Linha {row_number}: VALOR invalido: {text!r}") from exc
    return result.quantize(Decimal("0.01"))


def _safe_filename_part(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "-", value)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    return cleaned or "SEM INFORMACAO"
